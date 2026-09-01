# Las preguntas del levantamiento viven en tuniche-levantamiento-preguntas.py
# y de ahí salen las DOS salidas: el .xlsx que se manda y la página que se
# publica. Escritas dos veces, se habrían separado al primer cambio de
# redacción — y la que se separa siempre es la que el cliente termina leyendo.
import sys, math, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).parent))
from importlib import import_module
_p = import_module("tuniche-levantamiento-preguntas")
BLOQUES, STACK = _p.BLOQUES, _p.STACK
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation

# Marca adoOps, medida sobre el código de la propia web (dashboard360.css)
NAVY   = "FF0E1D33"
TEAL   = "FF0E8A82"
VERDE  = "FF2ED477"
SUAVE  = "FFE6F5F3"
AMBAR  = "FFB45309"
AMBARS = "FFFAF0E3"
TINTA  = "FF0B1523"
GRIS   = "FF43566A"
MUTED  = "FF7D8F9E"
LINEA  = "FFE4EAEE"
PLANO  = "FFF7F9FA"

borde = Border(bottom=Side(style="thin", color=LINEA))
wb = openpyxl.Workbook(); wb.remove(wb.active)

# ── Portada ──────────────────────────────────────────────────────────────────
ws = wb.create_sheet("adoOps"); ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 62
ws.column_dimensions["C"].width = 62

logo = XLImage("/Users/joaquintrujillo/Proyectos/adoOps.Digital/public/logo.png")
logo.height = 54; logo.width = int(54 * 553 / 155)
ws.add_image(logo, "B2")
ws.row_dimensions[2].height = 46

def linea(r, txt, col="B", font=None, alto=None, span=None):
    c = ws.cell(row=r, column={"B":2,"C":3}[col], value=txt)
    c.font = font or Font(name="Aptos", size=11, color=TINTA)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if alto: ws.row_dimensions[r].height = alto
    if span: ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=span)
    return c

linea(6,  "SEMILLAS TUNICHE · SISTEMA DE VISITAS A CAMPO", font=Font(name="Aptos", size=9, bold=True, color=TEAL))
linea(7,  "Levantamiento para costear", font=Font(name="Aptos", size=22, bold=True, color=TINTA), alto=30)
linea(8,  "SaaS o servidor propio", font=Font(name="Aptos", size=22, bold=True, color=TEAL), alto=30)
linea(10, "Necesitamos entender cómo funciona hoy su infraestructura y su operación para dimensionar "
          "bien las dos alternativas: mantener el sistema como servicio mensual, o instalarlo en sus "
          "servidores.", alto=46, span=3)

linea(12, "Cómo completarlo", font=Font(name="Aptos", size=13, bold=True, color=TINTA))
linea(13, "Tres hojas, una por tema. Las responde gente distinta y ninguna necesita esperar a las otras.",
      alto=18, span=3)
linea(14, "Escriban en la columna Respuesta. Si algo no lo saben todavía, déjenlo en blanco: es más útil "
          "que una aproximación.", alto=32, span=3)
linea(16, "Las preguntas marcadas «Necesario para costear» son las que pueden cambiar el alcance o el "
          "esfuerzo. Las demás ayudan a diseñar mejor, pero no bloquean una primera estimación.",
      alto=32, span=3)

linea(19, "Qué habría que definir si va a sus servidores", font=Font(name="Aptos", size=13, bold=True, color=TINTA))
for i, (hoy, adentro) in enumerate(STACK):
    r = 21 + i
    if i == 0:
        for col, t in ((2, "CÓMO FUNCIONA HOY"), (3, "QUÉ HABRÍA QUE DEFINIR")):
            c = ws.cell(row=20, column=col, value=t)
            c.font = Font(name="Aptos", size=9, bold=True, color="FFFFFFFF")
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[20].height = 20
    for col, t in ((2, hoy), (3, adentro)):
        c = ws.cell(row=r, column=col, value=t)
        c.font = Font(name="Aptos", size=10.5, color=TINTA if col == 2 else GRIS)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = borde
        if i % 2: c.fill = PatternFill("solid", fgColor=PLANO)
    ws.row_dimensions[r].height = 20

linea(30, "adoOps · agosto de 2026", font=Font(name="Aptos", size=9, color=MUTED))

def imprimible(hoja, repetir=None):
    """Apaisado y ajustado al ancho. Es un cuestionario: alguien lo va a
    imprimir para llevarlo a la reunión, y sin esto las columnas de la derecha
    —justo las de respuesta— se van a una segunda hoja."""
    hoja.page_setup.orientation = "landscape"
    hoja.sheet_properties.pageSetUpPr.fitToPage = True
    hoja.page_setup.fitToWidth = 1
    hoja.page_setup.fitToHeight = 0
    hoja.page_margins.left = hoja.page_margins.right = 0.4
    hoja.page_margins.top = hoja.page_margins.bottom = 0.5
    if repetir: hoja.print_title_rows = repetir
    hoja.oddFooter.right.text = "&P / &N"
    hoja.oddFooter.left.text = "adoOps · Semillas Tuniche"

imprimible(ws)

# ── Hojas de preguntas ───────────────────────────────────────────────────────
ENC = ["N°", "Sección", "Pregunta", "Ayuda", "Prioridad", "Respuesta", "Respondido por"]
ANCHOS = [5, 24, 58, 46, 20, 58, 20]

for titulo, quien, bajada, filas in BLOQUES:
    ws = wb.create_sheet(titulo[:31]); ws.sheet_view.showGridLines = False

    ws.cell(row=1, column=1, value=titulo).font = Font(name="Aptos", size=16, bold=True, color=TINTA)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.row_dimensions[1].height = 24

    c = ws.cell(row=2, column=1, value=quien.upper())
    c.font = Font(name="Aptos", size=9, bold=True, color=TEAL)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    c = ws.cell(row=3, column=1, value=bajada)
    c.font = Font(name="Aptos", size=10.5, color=GRIS)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)
    ws.row_dimensions[3].height = 30

    f = 5
    for i, t in enumerate(ENC, 1):
        c = ws.cell(row=f, column=i, value=t)
        c.font = Font(name="Aptos", size=9.5, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[f].height = 24

    for j, (seccion, pregunta, por_que, clave) in enumerate(filas):
        r = f + 1 + j
        vals = [j + 1, seccion, pregunta, por_que, "Necesario para costear" if clave else "Contexto", "", ""]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = borde
            c.alignment = Alignment(wrap_text=True, vertical="top",
                                    horizontal="center" if i == 1 else "left")
            if i == 1:   c.font = Font(name="Aptos", size=9.5, color=MUTED)
            elif i == 2: c.font = Font(name="Aptos", size=10, color=MUTED)
            elif i == 3: c.font = Font(name="Aptos", size=11, bold=clave, color=TINTA)
            elif i == 4: c.font = Font(name="Aptos", size=10, color=GRIS)
            elif i == 5: c.font = Font(name="Aptos", size=9, bold=clave, color=AMBAR if clave else MUTED)
            else:        c.font = Font(name="Aptos", size=11, color=TINTA)
            # La columna de respuesta se marca como editable con un fondo claro:
            # sin eso nadie sabe dónde escribir en una hoja de siete columnas.
            if i in (6, 7): c.fill = PatternFill("solid", fgColor="FFFFFFFF")
            elif clave:     c.fill = PatternFill("solid", fgColor=AMBARS)
            elif j % 2:     c.fill = PatternFill("solid", fgColor=PLANO)

        largo = max(len(pregunta), len(por_que))
        ws.row_dimensions[r].height = max(34, math.ceil(largo / 58) * 15 + 16)

    for i, a in enumerate(ANCHOS, 1):
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.freeze_panes = ws.cell(row=f + 1, column=3)
    ws.auto_filter.ref = f"A{f}:G{f + len(filas)}"
    # La cabecera se repite en cada hoja impresa: sin eso, a partir de la
    # segunda página nadie sabe cuál columna es «Respuesta».
    imprimible(ws, repetir=f"{f}:{f}")

wb.save("/Users/joaquintrujillo/Proyectos/adoOps.Digital/docs/sistema_funcional/Levantamiento para costear - adoOps.xlsx")
print("✓ hojas:", wb.sheetnames)
print("  preguntas:", sum(len(b[3]) for b in BLOQUES), "· marcadas:", sum(1 for b in BLOQUES for f in b[3] if f[3]))
